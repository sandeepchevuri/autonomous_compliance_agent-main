from openpyxl import Workbook
from fpdf import FPDF
from pathlib import Path

# Directories for reports
REPORT_DIR = Path("reports")
EXCEL_DIR = REPORT_DIR / "excel"
PDF_DIR = REPORT_DIR / "pdf"
FONT_DIR = Path("assets/fonts")
FONT_FILE = FONT_DIR / "DejaVuSans.ttf"

# Ensure folders exist
EXCEL_DIR.mkdir(parents=True, exist_ok=True)
PDF_DIR.mkdir(parents=True, exist_ok=True)

class ReportingAgent:
    def generate_excel(self, data: dict, filename: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Compliance Report"
        ws.append(["Category", "Value"])

        # Compliance data sections
        for key in ["obligations", "penalties", "entities", "dates"]:
            for item in data.get(key, []):
                ws.append([key.title(), item])

        # Compliance score
        ws.append(["Compliance Score", data.get("compliance_score", "N/A")])

        # Risk flags
        ws.append(["Risk Flags", ""])
        for flag in data.get("risk_flags", []):
            ws.append(["⚠️", flag])

        path = EXCEL_DIR / f"{filename}.xlsx"
        wb.save(path)
        return str(path)

    def generate_pdf(self, data: dict, filename: str) -> str:
        class CompliancePDF(FPDF):
            def header(self):
                self.set_font("DejaVu", "", 14)
                self.cell(0, 10, "Compliance Summary Report", ln=True, align="C")

            def add_section(self, title, items):
                self.set_font("DejaVu", "B", 12)
                self.cell(0, 10, title, ln=True)
                self.set_font("DejaVu", "", 11)
                for item in items:
                    self.multi_cell(0, 8, f"- {item}")
                self.ln(4)

        # Check font file
        if not FONT_FILE.exists():
            raise RuntimeError(f"Font file not found: {FONT_FILE}")

        pdf = CompliancePDF()
        pdf.add_font("DejaVu", "", str(FONT_FILE), uni=True)
        pdf.add_font("DejaVu", "B", str(FONT_FILE), uni=True)
        pdf.add_page()

        # Standard sections
        for key in ["obligations", "penalties", "entities", "dates"]:
            values = data.get(key, [])
            pdf.add_section(key.title(), values)

        # Compliance score
        pdf.add_section("Compliance Score", [str(data.get("compliance_score", "N/A"))])

        # Risk flags
        pdf.add_section("Risk Flags", data.get("risk_flags", []))

        out_path = PDF_DIR / f"{filename}.pdf"
        pdf.output(str(out_path))
        return str(out_path)
